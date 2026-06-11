"""TradeIQ TPO model evaluation — compare LLMs on the GenAI Sales Assistant flow.

For each candidate model it runs the real thin-orchestrator prompts:
  1. Intent classification (PROMPT-001) over the ground-truth dataset -> accuracy.
  2. Grounded response generation (PROMPT-002) from mock tool output -> judged on
     groundedness, relevance and format (LLM-as-judge), plus a deterministic
     figure-overlap groundedness signal.

Writes a comparison report so the best model(s) can be chosen.

Backend defaults to 'claude_code' (Claude Code subscription quota); use
'--backend providers' to compare metered provider models instead.

    uv run python scripts/evaluate_models.py --output results/model_eval.md
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import html
import json
import statistics
import sys
import time
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast

from langchain_core.language_models import BaseChatModel
from langchain_core.messages import HumanMessage, SystemMessage

from app.config import Settings
from app.eval import classified_intent, groundedness_overlap
from app.metrics import estimate_cost
from app.prompts import (
    INTENT_SYSTEM_PROMPT,
    INTENT_USER_TEMPLATE,
    INTENTS,
    RESPONSE_SYSTEM_PROMPT,
    RESPONSE_USER_TEMPLATE,
)
from app.tools import route_to_tool

# The evaluation always uses deterministic mock tool output (empty tool URLs),
# so groundedness scores stay reproducible and never depend on live services.
_EVAL_CFG = Settings(cdt_base_url="", erdc_base_url="")

INTENT_DATASET_PATH = Path(__file__).parent.parent / "tests" / "data" / "intent_dataset.json"

# Candidate models per backend.
CLAUDE_CODE_MODELS = ["opus", "sonnet", "haiku"]
PROVIDER_MODELS = [("openai", "gpt-4o"), ("anthropic", "claude-sonnet-4-6")]
# Azure AI Foundry candidate deployments. Listing many "keeps all models open":
# the eval auto-skips any that aren't deployed, so deploying more in Foundry
# lights them up here with no code change.
AZURE_MODELS = [
    "gpt-4o",
    "gpt-4o-mini",
    "gpt-4.1",
    "gpt-4.1-mini",
    "gpt-5",
    "gpt-5.1",
    "gpt-5-chat",
    "o4-mini",
    "model-router",
    "DeepSeek-V3",
    "Llama-3.3-70B-Instruct",
    "grok-3",
]

# Grounded response cases — each query + the params used to fetch its tool output.
GROUNDED_CASES: list[dict[str, Any]] = [
    {
        "query": "Why did you recommend the Easter display for Tesco?",
        "intent": "DATA_QUERY",
        "params": {"account": "Tesco"},
    },
    {"query": "Show me the recommendations for Carrefour.", "intent": "DATA_QUERY", "params": {"account": "Carrefour"}},
    {
        "query": "What are the best promo options for my remaining £80k budget?",
        "intent": "OPTIMIZER_RUN",
        "params": {"budget_remaining": 80000, "objective": "maximise ROI"},
    },
    {"query": "Give me options for a £40k budget.", "intent": "OPTIMIZER_RUN", "params": {"budget_remaining": 40000}},
]

# Agentic tool-selection cases — does the model itself choose the right tool?
# expected = the tool it should call, or None when it should call no tool
# (clarify / decline). Only run on the claude_code backend (Agent SDK).
TOOL_SELECTION_CASES: list[dict[str, Any]] = [
    {"query": "Why did you recommend the Easter display for Tesco?", "expected": "text_to_sql_lookup"},
    {"query": "List my top performing promos for Carrefour last quarter.", "expected": "text_to_sql_lookup"},
    {"query": "What are the best promo options for my remaining £80k budget at Tesco?", "expected": "optimizer_run"},
    {"query": "Re-optimise my Aldi plan to hit a 1.5x uplift guideline.", "expected": "optimizer_run"},
    {"query": "Show me the options.", "expected": None},
    {"query": "What's the weather in London today?", "expected": None},
]

JUDGE_PROMPT = """You are evaluating a TradeIQ Sales Assistant answer for a CPG commercial team.

The tool output below is the ONLY source of facts the answer may use:
{tool_output}

User query: "{query}"

Assistant answer:
{answer}

Score each dimension 1-5 (5 = best) and return ONLY a JSON object:
{{"groundedness": <1-5>, "relevance": <1-5>, "format": <1-5>, "notes": "<one sentence>"}}
- groundedness: every fact/number in the answer is supported by the tool output; NO fabrication.
- relevance: directly answers the user's query.
- format: concise business language, key figures in bold, a suggested next step."""

ACCOUNT_SCOPE = "All UK & EU grocery accounts"
PLANNING_PERIOD = "FY2025"

_CSS = """
body { font-family: system-ui, "Segoe UI", Arial, sans-serif; margin: 2rem; color: #1a1a1a; }
h1 { font-size: 1.4rem; margin-bottom: .25rem; }
h2 { font-size: 1.1rem; margin-top: 2rem; }
.meta { color: #666; margin-bottom: 1rem; }
table { border-collapse: collapse; width: 100%; margin-bottom: 1rem; }
th, td { border: 1px solid #ddd; padding: .4rem .6rem; text-align: right; vertical-align: top; }
th.l, td.l { text-align: left; }
thead { background: #0b5fff; color: #fff; }
tr.best { background: #e8f5e9; }
tr.bad { background: #fdecea; }
.rec { margin: 1rem 0; padding: .75rem; background: #e8f5e9; border-left: 4px solid #2e7d32; }
small { color: #888; }
.ans { max-width: 460px; white-space: pre-wrap; font-size: .85rem; text-align: left; }
pre { white-space: pre-wrap; font-size: .72rem; margin: 0; max-width: 380px; }
details summary { cursor: pointer; color: #0b5fff; }
svg text { fill: #1a1a1a; }
"""


@dataclass
class IntentRecord:
    """One intent-classification question and its result."""

    model: str
    query: str
    expected: str
    predicted: str
    correct: bool
    latency_ms: float
    cost_usd: float


@dataclass
class ResponseRecord:
    """One grounded-response question and its judged result."""

    model: str
    query: str
    intent: str
    tool: str
    answer: str
    groundedness: float
    relevance: float
    fmt: float
    figure_overlap: float
    latency_ms: float
    cost_usd: float
    params: str = "{}"
    tool_output: str = "{}"


@dataclass
class ToolSelectionRecord:
    """One agentic tool-selection question: which tool the model chose itself."""

    model: str
    query: str
    expected: str  # tool name, or "(no tool)"
    chose: str  # comma-joined tool names, or "(none)"
    correct: bool
    latency_ms: float
    cost_usd: float


@dataclass
class ModelScore:
    name: str
    intent_correct: int = 0
    intent_total: int = 0
    groundedness: list[float] = field(default_factory=list)
    relevance: list[float] = field(default_factory=list)
    fmt: list[float] = field(default_factory=list)
    figure_overlap: list[float] = field(default_factory=list)
    # Usage metrics (candidate-model calls only; judge calls excluded).
    latencies_ms: list[float] = field(default_factory=list)
    input_tokens: int = 0
    output_tokens: int = 0
    cost_usd: float = 0.0
    # Agentic tool-selection (claude_code backend only; may be 0/0 = not run).
    tool_selection_correct: int = 0
    tool_selection_total: int = 0
    # Per-question records for the data artifacts.
    intent_records: list[IntentRecord] = field(default_factory=list)
    response_records: list[ResponseRecord] = field(default_factory=list)
    tool_selection_records: list[ToolSelectionRecord] = field(default_factory=list)

    @property
    def intent_accuracy(self) -> float:
        return self.intent_correct / self.intent_total if self.intent_total else 0.0

    @property
    def tool_selection_accuracy(self) -> float:
        return self.tool_selection_correct / self.tool_selection_total if self.tool_selection_total else 0.0

    @staticmethod
    def _avg(xs: list[float]) -> float:
        return statistics.mean(xs) if xs else 0.0

    @property
    def avg_latency_ms(self) -> float:
        return self._avg(self.latencies_ms)

    @property
    def total_tokens(self) -> int:
        return self.input_tokens + self.output_tokens

    @property
    def composite(self) -> float:
        """Overall QUALITY score (0-1): intent accuracy + judged dims + figure overlap."""
        judged = self._avg(self.groundedness + self.relevance + self.fmt) / 5
        return statistics.mean([self.intent_accuracy, judged, self._avg(self.figure_overlap)])

    @property
    def value_index(self) -> float:
        """Composite quality per US dollar of estimated cost. Higher = better value."""
        return round(self.composite / max(self.cost_usd, 1e-6), 1)

    def record_call(self, latency_ms: float, input_tokens: int, output_tokens: int) -> None:
        """Record usage metrics for one candidate-model call."""
        self.latencies_ms.append(latency_ms)
        self.input_tokens += input_tokens
        self.output_tokens += output_tokens
        self.cost_usd = round(self.cost_usd + estimate_cost(self.name, input_tokens, output_tokens), 6)


def _parse_judge(text: str) -> dict[str, float]:
    try:
        start = text.index("{")
        data = json.loads(text[start : text.rindex("}") + 1])
    except (ValueError, json.JSONDecodeError):
        return {}
    out: dict[str, float] = {}
    for key in ("groundedness", "relevance", "format"):
        try:
            out[key] = float(data[key])
        except (KeyError, TypeError, ValueError):
            out[key] = 0.0
    return out


async def _complete(llm: BaseChatModel, system: str, user: str) -> tuple[str, float, int, int]:
    """Return (text, latency_ms, input_tokens, output_tokens) for one call."""
    start = time.perf_counter()
    resp = await llm.ainvoke([SystemMessage(content=system), HumanMessage(content=user)])
    latency_ms = round((time.perf_counter() - start) * 1000, 1)
    usage = getattr(resp, "usage_metadata", None)
    in_tok = int(usage.get("input_tokens", 0)) if isinstance(usage, dict) else 0
    out_tok = int(usage.get("output_tokens", 0)) if isinstance(usage, dict) else 0
    return str(resp.content), latency_ms, in_tok, out_tok


async def _evaluate_model(
    name: str,
    llm: BaseChatModel,
    judge: BaseChatModel,
    intent_cases: list[dict[str, str]],
) -> ModelScore:
    score = ModelScore(name=name)

    # 1. Intent classification accuracy (per-question records + usage).
    for case in intent_cases:
        user = INTENT_USER_TEMPLATE.format(
            query=case["query"], history="[]", account_scope=ACCOUNT_SCOPE, planning_period=PLANNING_PERIOD
        )
        try:
            text, latency_ms, in_tok, out_tok = await _complete(llm, INTENT_SYSTEM_PROMPT, user)
        except Exception as exc:
            print(f"  intent ERROR ({case['query'][:40]}): {exc}")
            continue
        score.intent_total += 1
        score.record_call(latency_ms, in_tok, out_tok)
        predicted = classified_intent(text) or "(unparsed)"
        correct = predicted == case["expected"]
        if correct:
            score.intent_correct += 1
        score.intent_records.append(
            IntentRecord(
                model=name,
                query=case["query"],
                expected=case["expected"],
                predicted=predicted,
                correct=correct,
                latency_ms=latency_ms,
                cost_usd=round(estimate_cost(name, in_tok, out_tok), 6),
            )
        )
    print(f"  intent accuracy: {score.intent_accuracy * 100:.1f}% ({score.intent_correct}/{score.intent_total})")

    # 2. Grounded response generation, judged (per-question records + usage).
    for case in GROUNDED_CASES:
        tool_name, tool_desc, tool_output = await route_to_tool(
            str(case["intent"]), dict(case["params"]), _EVAL_CFG, query=str(case["query"])
        )
        user = RESPONSE_USER_TEMPLATE.format(
            query=case["query"],
            intent=case["intent"],
            tool_name=tool_name,
            tool_description=tool_desc,
            tool_output=json.dumps(tool_output, indent=2),
            account_scope=ACCOUNT_SCOPE,
            planning_period=PLANNING_PERIOD,
        )
        try:
            answer, latency_ms, in_tok, out_tok = await _complete(llm, RESPONSE_SYSTEM_PROMPT, user)
        except Exception as exc:
            print(f"  response ERROR ({case['query'][:40]}): {exc}")
            continue

        score.record_call(latency_ms, in_tok, out_tok)
        overlap = groundedness_overlap(answer, tool_output)
        score.figure_overlap.append(overlap)
        judge_text, *_ = await _complete(
            judge,
            "You are a strict, fair evaluator. Return only JSON.",
            JUDGE_PROMPT.format(tool_output=json.dumps(tool_output, indent=2), query=case["query"], answer=answer),
        )
        judged = _parse_judge(judge_text)
        g, r, f = judged.get("groundedness", 0.0), judged.get("relevance", 0.0), judged.get("format", 0.0)
        score.groundedness.append(g)
        score.relevance.append(r)
        score.fmt.append(f)
        score.response_records.append(
            ResponseRecord(
                model=name,
                query=str(case["query"]),
                intent=str(case["intent"]),
                tool=tool_name,
                answer=answer,
                groundedness=g,
                relevance=r,
                fmt=f,
                figure_overlap=overlap,
                latency_ms=latency_ms,
                cost_usd=round(estimate_cost(name, in_tok, out_tok), 6),
                params=json.dumps(dict(case["params"])),
                tool_output=json.dumps(tool_output),
            )
        )
    print(
        f"  groundedness: judge={score._avg(score.groundedness):.1f}/5 "
        f"figure-overlap={score._avg(score.figure_overlap) * 100:.0f}% "
        f"| avg latency {score.avg_latency_ms:.0f}ms · cost ${score.cost_usd:.4f}"
    )
    return score


async def _evaluate_tool_selection(
    name: str, cases: list[dict[str, Any]], agent: Any
) -> tuple[int, int, list[ToolSelectionRecord]]:
    """Run an AGENTIC orchestrator over labelled cases — does the model pick the
    right tool itself? ``agent`` is an AgenticOrchestrator (Claude Agent SDK) or a
    ToolCallingOrchestrator (native bind_tools, e.g. Azure/OpenAI); both expose
    ``.run(query)`` returning ``.tool_names`` + ``.metrics``."""
    correct = 0
    records: list[ToolSelectionRecord] = []
    for case in cases:
        expected = case["expected"]
        try:
            result = await agent.run(str(case["query"]))
        except Exception as exc:
            print(f"  tool-selection ERROR ({str(case['query'])[:40]}): {exc}")
            continue
        chose = result.tool_names
        ok = len(chose) == 0 if expected is None else expected in chose
        correct += ok
        records.append(
            ToolSelectionRecord(
                model=name,
                query=str(case["query"]),
                expected=str(expected) if expected else "(no tool)",
                chose=", ".join(chose) or "(none)",
                correct=bool(ok),
                latency_ms=result.metrics.latency_ms,
                cost_usd=result.metrics.cost_usd,
            )
        )
    total = len(records)
    print(f"  tool-selection accuracy: {correct}/{total}" + (f" ({correct / total * 100:.0f}%)" if total else ""))
    return correct, total, records


def _fmt_tool_selection(s: ModelScore) -> str:
    """Tool-selection accuracy as a percentage, or 'n/a' when it wasn't run."""
    return f"{s.tool_selection_accuracy * 100:.0f}%" if s.tool_selection_total else "n/a"


def _render_report(scores: list[ModelScore], intent_n: int, timestamp: str, scope: str) -> str:
    lines = [
        "# TradeIQ TPO — Model Evaluation",
        "",
        f"Generated: {timestamp} · Intent cases: {intent_n} · Grounded response cases: {len(GROUNDED_CASES)}",
        "",
        f"_{scope}_",
        "",
        "| Model | Intent acc. | Tool sel. | Groundedness /5 | Figure overlap | Relevance /5 | Format /5 "
        "| Avg latency (ms) | Cost (USD) | Tokens | Value | Composite |",
        "|---|---|---|---|---|---|---|---|---|---|---|---|",
    ]
    for s in sorted(scores, key=lambda x: -x.composite):
        lines.append(
            f"| **{s.name}** | {s.intent_accuracy * 100:.1f}% | {_fmt_tool_selection(s)} "
            f"| {s._avg(s.groundedness):.1f} | {s._avg(s.figure_overlap) * 100:.0f}% "
            f"| {s._avg(s.relevance):.1f} | {s._avg(s.fmt):.1f} "
            f"| {s.avg_latency_ms:.0f} | ${s.cost_usd:.4f} | {s.total_tokens} "
            f"| {s.value_index} | {s.composite:.3f} |"
        )
    if scores:
        best = max(scores, key=lambda x: x.composite)
        lines += ["", f"**Recommended:** {best.name} (composite {best.composite:.3f})."]
    return "\n".join(lines) + "\n"


def _confusion(
    scores: list[ModelScore],
) -> tuple[list[str], dict[tuple[str, str], int], dict[str, list[int]]]:
    """Aggregate confusion matrix + per-intent accuracy across all models."""
    labels = [*INTENTS, "(unparsed)"]
    matrix: dict[tuple[str, str], int] = {}
    per_intent: dict[str, list[int]] = {i: [0, 0] for i in INTENTS}
    for s in scores:
        for rec in s.intent_records:
            key = (rec.expected, rec.predicted)
            matrix[key] = matrix.get(key, 0) + 1
            if rec.expected in per_intent:
                per_intent[rec.expected][1] += 1
                if rec.correct:
                    per_intent[rec.expected][0] += 1
    return labels, matrix, per_intent


def _svg_bars(ranked: list[ModelScore]) -> str:
    """Inline SVG horizontal bar chart of composite scores (no JS)."""
    bar_h, gap, width = 22, 8, 360
    parts: list[str] = []
    y = 0
    for s in ranked:
        bar_w = int(width * min(max(s.composite, 0.0), 1.0))
        parts.append(
            f'<text x="0" y="{y + 15}" font-size="12">{html.escape(s.name)}</text>'
            f'<rect x="90" y="{y}" width="{bar_w}" height="{bar_h}" rx="2" fill="#0b5fff"></rect>'
            f'<text x="{95 + bar_w}" y="{y + 15}" font-size="11">{s.composite:.3f}</text>'
        )
        y += bar_h + gap
    return f'<svg width="520" height="{max(y, 1)}" role="img" aria-label="composite scores">{"".join(parts)}</svg>'


def _render_html(scores: list[ModelScore], intent_n: int, timestamp: str, scope: str) -> str:
    ranked = sorted(scores, key=lambda x: -x.composite)
    if not ranked:
        return "<!doctype html><html><body><p>No results.</p></body></html>\n"
    best = ranked[0]

    summary = "".join(
        f'<tr class="{"best" if s.name == best.name else ""}">'
        f'<td class="l">{html.escape(s.name)}</td>'
        f"<td>{s.intent_accuracy * 100:.1f}%</td><td>{_fmt_tool_selection(s)}</td>"
        f"<td>{s._avg(s.groundedness):.1f}</td>"
        f"<td>{s._avg(s.figure_overlap) * 100:.0f}%</td><td>{s._avg(s.relevance):.1f}</td>"
        f"<td>{s._avg(s.fmt):.1f}</td><td>{s.avg_latency_ms:.0f}</td><td>${s.cost_usd:.4f}</td>"
        f"<td>{s.total_tokens}</td><td>{s.value_index}</td><td><strong>{s.composite:.3f}</strong></td></tr>"
        for s in ranked
    )
    intent_rows = "".join(
        f'<tr class="{"" if rec.correct else "bad"}"><td class="l">{html.escape(rec.model)}</td>'
        f'<td class="l">{html.escape(rec.query)}</td><td>{rec.expected}</td><td>{rec.predicted}</td>'
        f"<td>{'PASS' if rec.correct else 'FAIL'}</td><td>{rec.latency_ms:.0f}</td></tr>"
        for s in ranked
        for rec in s.intent_records
    )
    response_rows = "".join(
        f'<tr><td class="l">{html.escape(rec.model)}</td><td class="l">{html.escape(rec.query)}</td>'
        f"<td>{rec.intent}</td><td>{html.escape(rec.tool)}</td><td>{rec.groundedness:.0f}</td>"
        f"<td>{rec.relevance:.0f}</td><td>{rec.fmt:.0f}</td><td>{rec.figure_overlap * 100:.0f}%</td>"
        f'<td class="ans">{html.escape(rec.answer)}</td>'
        f'<td class="l"><details><summary>data</summary><pre>{html.escape(rec.tool_output)}</pre></details></td></tr>'
        for s in ranked
        for rec in s.response_records
    )
    tool_selection_rows = "".join(
        f'<tr class="{"" if rec.correct else "bad"}"><td class="l">{html.escape(rec.model)}</td>'
        f'<td class="l">{html.escape(rec.query)}</td><td>{html.escape(rec.expected)}</td>'
        f"<td>{html.escape(rec.chose)}</td><td>{'PASS' if rec.correct else 'FAIL'}</td>"
        f"<td>{rec.latency_ms:.0f}</td></tr>"
        for s in ranked
        for rec in s.tool_selection_records
    )

    labels, matrix, per_intent = _confusion(scores)
    conf_head = "".join(f"<th>{p}</th>" for p in labels)
    conf_body = ""
    for expected in INTENTS:
        cells = ""
        for predicted in labels:
            n = matrix.get((expected, predicted), 0)
            cls = "best" if (expected == predicted and n) else ("bad" if (expected != predicted and n) else "")
            cells += f'<td class="{cls}">{n or ""}</td>'
        correct, total = per_intent[expected]
        conf_body += (
            f'<tr><td class="l">{expected}</td>{cells}<td>{(correct / total * 100) if total else 0:.0f}%</td></tr>'
        )

    tool_selection_section = (
        (
            "<h2>Agentic tool selection — per question (the model decides the tool)</h2>"
            '<table><thead><tr><th class="l">Model</th><th class="l">Question</th><th>Expected</th>'
            "<th>Chose</th><th>Result</th><th>Latency (ms)</th></tr></thead>"
            f"<tbody>{tool_selection_rows}</tbody></table>"
        )
        if tool_selection_rows
        else ""
    )

    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>TradeIQ TPO — Model Evaluation</title>
<style>{_CSS}</style></head><body>
<h1>TradeIQ TPO — Model Evaluation</h1>
<div class="meta">Generated: {timestamp} &middot; Intent cases: {intent_n}
&middot; Grounded cases: {len(GROUNDED_CASES)}</div>
<div class="meta"><strong>Scope:</strong> {html.escape(scope)}</div>
<div class="rec"><strong>Recommended:</strong> {html.escape(best.name)} (composite {best.composite:.3f})</div>
<h2>Composite score</h2>
{_svg_bars(ranked)}
<h2>Summary</h2>
<table><thead><tr><th class="l">Model</th><th>Intent acc.</th><th>Tool sel.</th><th>Groundedness /5</th>
<th>Figure overlap</th><th>Relevance /5</th><th>Format /5</th><th>Avg latency (ms)</th>
<th>Cost (USD)</th><th>Tokens</th><th>Value</th><th>Composite</th></tr></thead>
<tbody>{summary}</tbody></table>
<h2>Intent confusion matrix (row = expected, col = predicted)</h2>
<table><thead><tr><th class="l">expected \\ predicted</th>{conf_head}<th>Accuracy</th></tr></thead>
<tbody>{conf_body}</tbody></table>
<h2>Intent classification — per question</h2>
<table><thead><tr><th class="l">Model</th><th class="l">Question</th><th>Expected</th>
<th>Predicted</th><th>Result</th><th>Latency (ms)</th></tr></thead>
<tbody>{intent_rows}</tbody></table>
<h2>Grounded responses — per question</h2>
<table><thead><tr><th class="l">Model</th><th class="l">Question</th><th>Intent</th><th>Tool</th>
<th>Grounded /5</th><th>Relevance /5</th><th>Format /5</th><th>Figure overlap</th>
<th class="l">Answer</th><th class="l">Tool output</th></tr></thead>
<tbody>{response_rows}</tbody></table>
{tool_selection_section}
<p><small>Value = composite quality per US dollar of estimated cost. Groundedness = answer uses only tool
output (LLM judge + figure-overlap check). Tool selection = the agentic orchestrator (Claude Agent SDK)
deciding the tool itself. Claude Code usage is subscription-quota billed.</small></p>
</body></html>
"""


_HISTORY_HEADER = [
    "timestamp",
    "model",
    "intent_acc",
    "tool_selection_acc",
    "groundedness",
    "relevance",
    "format",
    "figure_overlap",
    "avg_latency_ms",
    "cost_usd",
    "value_index",
    "composite",
]


def _append_history(scores: list[ModelScore], timestamp: str, path: Path) -> list[list[Any]]:
    """Append this round to the cumulative history CSV, rewriting the whole file
    with the current header. Rows written under an older schema are migrated by
    column name (missing columns become empty), so the CSV stays consistent even
    as the dimensions evolve."""
    prior: list[list[str]] = []
    if path.exists():
        with path.open(encoding="utf-8") as fh:
            existing = list(csv.reader(fh))
        if existing:
            old_header = existing[0]
            for row in existing[1:]:
                mapped = dict(zip(old_header, row, strict=False))
                prior.append([mapped.get(col, "") for col in _HISTORY_HEADER])

    new_rows = [
        [
            timestamp,
            s.name,
            round(s.intent_accuracy * 100, 1),
            round(s.tool_selection_accuracy * 100, 0) if s.tool_selection_total else "",
            round(s._avg(s.groundedness), 2),
            round(s._avg(s.relevance), 2),
            round(s._avg(s.fmt), 2),
            round(s._avg(s.figure_overlap) * 100, 0),
            round(s.avg_latency_ms, 0),
            round(s.cost_usd, 6),
            s.value_index,
            round(s.composite, 3),
        ]
        for s in scores
    ]

    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(_HISTORY_HEADER)
        writer.writerows(prior + new_rows)
    return [_HISTORY_HEADER, *prior, *new_rows]


def _export_json(scores: list[ModelScore], intent_n: int, timestamp: str, backend: str) -> str:
    """Machine-readable artifact: summary + every per-question record."""
    payload = {
        "timestamp": timestamp,
        "backend": backend,
        "intent_cases": intent_n,
        "grounded_cases": len(GROUNDED_CASES),
        "models": [
            {
                "model": s.name,
                "intent_accuracy": round(s.intent_accuracy, 4),
                "tool_selection_accuracy": (round(s.tool_selection_accuracy, 4) if s.tool_selection_total else None),
                "groundedness": round(s._avg(s.groundedness), 3),
                "relevance": round(s._avg(s.relevance), 3),
                "format": round(s._avg(s.fmt), 3),
                "figure_overlap": round(s._avg(s.figure_overlap), 3),
                "avg_latency_ms": s.avg_latency_ms,
                "cost_usd": s.cost_usd,
                "input_tokens": s.input_tokens,
                "output_tokens": s.output_tokens,
                "value_index": s.value_index,
                "composite": round(s.composite, 4),
                "intent_records": [asdict(r) for r in s.intent_records],
                "response_records": [asdict(r) for r in s.response_records],
                "tool_selection_records": [asdict(r) for r in s.tool_selection_records],
            }
            for s in scores
        ],
    }
    return json.dumps(payload, indent=2)


def _write_xlsx(scores: list[ModelScore], path: Path, history_rows: list[list[str]]) -> None:
    """Workbook: Summary, Intent, Responses, Confusion, Trend — with conditional formatting."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    bold = Font(bold=True)
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    score_scale = ColorScaleRule(
        start_type="num",
        start_value=1,
        start_color="FFC7CE",
        mid_type="num",
        mid_value=3,
        mid_color="FFEB9C",
        end_type="num",
        end_value=5,
        end_color="C6EFCE",
    )

    def _fill(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        for row in rows:
            ws.append(row)

    summary = wb.active
    summary.title = "Summary"
    _fill(
        summary,
        [
            "Model",
            "Intent acc %",
            "Tool sel %",
            "Groundedness /5",
            "Figure overlap %",
            "Relevance /5",
            "Format /5",
            "Avg latency ms",
            "Cost USD",
            "Input tokens",
            "Output tokens",
            "Value index",
            "Composite",
        ],
        [
            [
                s.name,
                round(s.intent_accuracy * 100, 1),
                round(s.tool_selection_accuracy * 100, 0) if s.tool_selection_total else "n/a",
                round(s._avg(s.groundedness), 2),
                round(s._avg(s.figure_overlap) * 100, 0),
                round(s._avg(s.relevance), 2),
                round(s._avg(s.fmt), 2),
                round(s.avg_latency_ms, 0),
                round(s.cost_usd, 6),
                s.input_tokens,
                s.output_tokens,
                s.value_index,
                round(s.composite, 3),
            ]
            for s in sorted(scores, key=lambda x: -x.composite)
        ],
    )

    intent_rows = [
        [
            r.model,
            r.query,
            r.expected,
            r.predicted,
            "PASS" if r.correct else "FAIL",
            round(r.latency_ms, 0),
            round(r.cost_usd, 6),
        ]
        for s in scores
        for r in s.intent_records
    ]
    intent_ws = wb.create_sheet("Intent")
    _fill(
        intent_ws,
        ["Model", "Question", "Expected intent", "Predicted intent", "Result", "Latency ms", "Cost USD"],
        intent_rows,
    )
    if intent_rows:
        rng = f"E2:E{len(intent_rows) + 1}"
        intent_ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PASS"'], fill=green))
        intent_ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"FAIL"'], fill=red))

    resp_rows = [
        [
            r.model,
            r.query,
            r.intent,
            r.tool,
            r.groundedness,
            r.relevance,
            r.fmt,
            round(r.figure_overlap * 100, 0),
            round(r.latency_ms, 0),
            round(r.cost_usd, 6),
            r.answer,
            r.params,
            r.tool_output,
        ]
        for s in scores
        for r in s.response_records
    ]
    resp_ws = wb.create_sheet("Responses")
    _fill(
        resp_ws,
        [
            "Model",
            "Question",
            "Intent",
            "Tool",
            "Groundedness /5",
            "Relevance /5",
            "Format /5",
            "Figure overlap %",
            "Latency ms",
            "Cost USD",
            "Answer",
            "Tool params",
            "Tool output",
        ],
        resp_rows,
    )
    if resp_rows:
        resp_ws.conditional_formatting.add(f"E2:G{len(resp_rows) + 1}", score_scale)

    labels, matrix, per_intent = _confusion(scores)
    confusion = wb.create_sheet("Confusion")
    _fill(
        confusion,
        ["expected \\ predicted", *labels, "Accuracy %"],
        [
            [
                e,
                *[matrix.get((e, p), 0) for p in labels],
                round(per_intent[e][0] / per_intent[e][1] * 100, 0) if per_intent[e][1] else 0,
            ]
            for e in INTENTS
        ],
    )

    ts_rows = [
        [
            r.model,
            r.query,
            r.expected,
            r.chose,
            "PASS" if r.correct else "FAIL",
            round(r.latency_ms, 0),
            round(r.cost_usd, 6),
        ]
        for s in scores
        for r in s.tool_selection_records
    ]
    if ts_rows:
        ts_ws = wb.create_sheet("ToolSelection")
        _fill(ts_ws, ["Model", "Question", "Expected tool", "Chose", "Result", "Latency ms", "Cost USD"], ts_rows)
        rng = f"E2:E{len(ts_rows) + 1}"
        ts_ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PASS"'], fill=green))
        ts_ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"FAIL"'], fill=red))

    trend = wb.create_sheet("Trend")
    for row in history_rows:
        trend.append(row)
    if trend[1]:
        for cell in trend[1]:
            cell.font = bold

    wb.save(str(path))


async def _azure_live_models(candidates: list[str]) -> list[str]:
    """Probe each candidate Azure deployment; keep the ones that actually answer.
    This is how the eval 'keeps all models open' — undeployed names are skipped."""
    from app.llm_factory import build_llm

    live: list[str] = []
    for name in candidates:
        try:
            await build_llm(model=name, provider="azure").ainvoke("ping")
            live.append(name)
            print(f"  azure deployed: {name}")
        except Exception as exc:
            print(f"  azure skip {name}: {str(exc).splitlines()[0][:80]}")
    return live


def _build_ts_agent(backend: str, name: str, llm: BaseChatModel) -> Any:
    """Pick the agentic orchestrator for tool-selection scoring on this backend."""
    if backend == "claude_code":
        from app.agents.agentic_orchestrator import AgenticOrchestrator

        return AgenticOrchestrator(model=name, cfg=_EVAL_CFG)
    from app.agents.tool_calling_orchestrator import ToolCallingOrchestrator

    return ToolCallingOrchestrator(llm=llm, cfg=_EVAL_CFG)


async def run_evaluation(
    intent_sample: int, output_path: Path, backend: str, tool_selection_sample: int = len(TOOL_SELECTION_CASES)
) -> None:
    from app.llm_factory import build_llm

    timestamp = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    slug = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    dataset: list[dict[str, str]] = json.loads(INTENT_DATASET_PATH.read_text(encoding="utf-8"))
    # The dataset is grouped by intent, so stride-sample to span all intents
    # even for a small --intent-sample.
    if intent_sample > 0:
        stride = max(1, len(dataset) // intent_sample)
        intent_cases = dataset[::stride][:intent_sample]
    else:
        intent_cases = dataset

    # Tool-selection runs on every backend (Agent SDK for claude_code; native
    # bind_tools for azure / providers).
    ts_cases = TOOL_SELECTION_CASES[:tool_selection_sample] if tool_selection_sample > 0 else []
    run_ts = bool(ts_cases)
    print(
        f"Intent cases: {len(intent_cases)} | Grounded cases: {len(GROUNDED_CASES)} | "
        f"Tool-selection cases: {len(ts_cases) if run_ts else 0} | backend: {backend}"
    )

    if backend == "claude_code":
        models = [(m, build_llm(model=m, provider="claude_code")) for m in CLAUDE_CODE_MODELS]
        judge = build_llm(model="haiku", provider="claude_code")
    elif backend == "azure":
        live = await _azure_live_models(AZURE_MODELS)
        if not live:
            print("No Azure deployments reachable — aborting.")
            return
        models = [(m, build_llm(model=m, provider="azure")) for m in live]
        judge = build_llm(model=live[0], provider="azure")  # a deployed model judges
    else:  # providers (metered API keys)
        models = [(name, build_llm(model=name, provider=prov)) for prov, name in PROVIDER_MODELS]
        judge = build_llm(model="gpt-4o-mini", provider="openai")

    scores: list[ModelScore] = []
    for name, llm in models:
        print(f"\n▶ {name}")
        score = await _evaluate_model(name, llm, judge, intent_cases)
        if run_ts:
            agent = _build_ts_agent(backend, name, llm)
            correct, total, records = await _evaluate_tool_selection(name, ts_cases, agent)
            score.tool_selection_correct, score.tool_selection_total = correct, total
            score.tool_selection_records = records
        scores.append(score)

    names = ", ".join(s.name for s in scores)
    ts_note = f" Tool-selection measured on {len(ts_cases)} cases." if run_ts else ""
    if backend == "claude_code":
        scope = (
            f"Claude family only ({names}), via the Claude Code subscription quota. "
            "Other providers need their own backend (`--backend azure` / `providers`)." + ts_note
        )
    elif backend == "azure":
        scope = (
            f"Azure AI Foundry, OpenAI-compatible ({names}). Undeployed catalog models "
            "were auto-skipped — deploy more in Foundry to include them." + ts_note
        )
    else:
        scope = f"Metered provider APIs ({backend}): {names}." + ts_note

    n = len(intent_cases)
    stem = output_path.with_suffix("")  # e.g. results/model_eval
    stem.parent.mkdir(parents=True, exist_ok=True)
    report = _render_report(scores, n, timestamp, scope)
    html_doc = _render_html(scores, n, timestamp, scope)
    json_doc = _export_json(scores, n, timestamp, backend)
    history = _append_history(scores, timestamp, stem.with_name("model_eval_history.csv"))

    # Write a timestamped set (history) and a "latest" set (easy to open).
    for tag in (slug, "latest"):
        stem.with_name(f"{stem.name}_{tag}.md").write_text(report, encoding="utf-8")
        stem.with_name(f"{stem.name}_{tag}.html").write_text(html_doc, encoding="utf-8")
        stem.with_name(f"{stem.name}_{tag}.json").write_text(json_doc, encoding="utf-8")
        _write_xlsx(scores, stem.with_name(f"{stem.name}_{tag}.xlsx"), history)

    print(f"\n✅ Round {slug} artifacts (.md/.html/.xlsx/.json + history.csv) in {stem.parent}\n")
    print(report)


def main() -> None:
    # Windows consoles default to cp1252 and choke on the report's unicode
    # (£, ▶, ·, …). Force UTF-8 output so the eval runs everywhere.
    cast(Any, sys.stdout).reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser(description="TradeIQ TPO model evaluation")
    parser.add_argument("--intent-sample", type=int, default=24, help="Number of intent cases to test (0 = all)")
    parser.add_argument("--output", type=Path, default=Path("results/model_eval.md"), help="Output markdown file")
    parser.add_argument("--backend", choices=["claude_code", "azure", "providers"], default="claude_code")
    parser.add_argument(
        "--tool-selection-sample",
        type=int,
        default=len(TOOL_SELECTION_CASES),
        help="Agentic tool-selection cases per model (claude_code only; 0 = skip)",
    )
    parser.add_argument(
        "--no-tool-selection", action="store_true", help="Skip the agentic tool-selection dimension entirely"
    )
    args = parser.parse_args()
    ts_sample = 0 if args.no_tool_selection else args.tool_selection_sample
    asyncio.run(run_evaluation(args.intent_sample, args.output, args.backend, ts_sample))


if __name__ == "__main__":
    main()
